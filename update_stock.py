import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import sys

# Protected configuration
_K = "pasindu6861534"
_W = r"D:\Antigravity\Stock Cal\Today_Stocks.xlsx"
_T = r"D:\Antigravity\Stock Cal\Yesterday_Stocks.xlsx"
_O = r"D:\Antigravity\Stock Cal\Updated_Stock_Template.xlsx"

# Formatting tokens
_F1 = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
_F2 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_F3 = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

def _exec_logic():
    print("--- Running Process ---")

    if not os.path.exists(_W):
        return

    try:
        _d1 = pd.read_excel(_W)
        _m1 = dict(zip(_d1['SKU'].astype(str).str.strip(), _d1['Availability']))
    except:
        return

    if not os.path.exists(_T):
        return

    try:
        _wb = load_workbook(_T)
        _ws = _wb.active 
    except:
        return

    if _ws.tables:
        _tn = list(_ws.tables.keys())
        for _n in _tn:
            del _ws.tables[_n]

    _c1 = 0
    for _r in _ws.iter_rows(min_row=2):
        _s_cell = _r[0]   
        _q_cell = _r[1] 
        
        _s = str(_s_cell.value).strip() if _s_cell.value else None
        
        if _s and _s in _m1:
            _nv = _m1[_s]
            if pd.isna(_nv): continue
            
            try:
                _nv = int(_nv)
            except:
                continue

            _q_cell.value = _nv
            _q_cell.number_format = '0'
            _c1 += 1
            
            if _nv <= 7:
                _q_cell.fill = _F3
            elif 8 <= _nv <= 10:
                _q_cell.fill = _F2
            else:
                _q_cell.fill = _F1
                
    try:
        _wb.save(_O)
        print(f"Process complete. Output generated.")
    except:
        print("Save failed.")

if __name__ == "__main__":
    _exec_logic()
