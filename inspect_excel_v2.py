import pandas as pd
import os

files = {
    "MSKU": r"D:\Antigravity\Stock Cal\MSKU Sheets.xlsx",
    "Tiktok": r"D:\Antigravity\Stock Cal\Tiktok Template.xlsx",
    "Warehouse": r"D:\Antigravity\Stock Cal\Warehouse sheet.xlsx"
}

print("--- Investigating MSKU Sheets.xlsx ---")
# Try reading as CSV first just in case
try:
    with open(files["MSKU"], 'r', errors='ignore') as f:
        head = f.read(100)
        print(f"First 100 bytes identifying: {head}")
except Exception as e:
    print(f"Could not read as text: {e}")

# Try reading as XLS
try:
    df = pd.read_excel(files["MSKU"], engine='xlrd')
    print("Read with xlrd:")
    print(df.columns.tolist())
except Exception as e:
    print(f"xlrd failed: {e}")

print("\n--- Investigating Tiktok Template.xlsx ---")
# Try reading with openpyxl but catching errors or maybe just standard pandas again but purely data? 
# Pandas read_excel doesn't have a 'ignore_styles' easily, but openpyxl does.
try:
    from openpyxl import load_workbook
    wb = load_workbook(files["Tiktok"], data_only=True, read_only=True)
    ws = wb.active
    print("Successfully opened with openpyxl read_only=True")
    # Print first 5 rows
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 5: break
        print(row)
except Exception as e:
    print(f"openpyxl failed: {e}")


print("\n--- Investigating Warehouse sheet.xlsx ---")
try:
    # Read more rows to find the header
    df = pd.read_excel(files["Warehouse"], header=None, nrows=20)
    print(df)
except Exception as e:
    print(f"Error: {e}")
